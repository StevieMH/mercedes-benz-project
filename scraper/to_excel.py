"""
to_excel.py
───────────
Converts Mercedes-Benz scraped JSON into a clean Excel spreadsheet.
Handles BOTH used and new car data with their correct fields.

HOW TO USE:
  For used cars:  python to_excel.py used
  For new cars:   python to_excel.py new

Requirements:
  pip install openpyxl
"""

import json
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── ROUTING ──────────────────────────────────────────────────────────────────
mode = sys.argv[1].lower() if len(sys.argv) > 1 else "used"

if mode == "new":
    INPUT_FILE = "data/new/mercedes_new_cars_FULL.json"
    OUTPUT_FILE = "output/mercedes_new_vehicles.xlsx"
    SHEET_LABEL = "New Vehicles"
else:
    INPUT_FILE = "data/used/mercedes_used_cars_FULL.json"
    OUTPUT_FILE = "output/mercedes_used_vehicles.xlsx"
    SHEET_LABEL = "Used Vehicles"
# ──────────────────────────────────────────────────────────────────────────────


def extract_used(v: dict) -> dict:
    engine = v.get("Engine", {})
    retailer = v.get("Retailer", {})
    source = v.get("VehicleSource", {})
    brand = v.get("Brand", {})

    return {
        # Identity
        "Listing ID":            v.get("Id"),
        "VIN":                   v.get("Vin"),
        "Registration":          v.get("RegistrationNumber"),
        "Registration Date":     v.get("RegistrationDate"),
        # Model
        "Brand":                 brand.get("Description", v.get("VehicleClass", "")),
        "Model":                 v.get("Model"),
        "Variant":               v.get("Description"),
        "Trim":                  v.get("ShortDescription"),
        "Body Style":            v.get("BodyStyle"),
        "Marketing Code":        v.get("MarketingCode"),
        # Pricing
        "Price (£)":             v.get("ActualPrice", 0),
        "Total Offer Value (£)": v.get("TotalOfferValue", 0),
        "Has Offer":             "Yes" if v.get("HasPromotionalOffer") else "No",
        "Campaign Offer (£)":    v.get("CampaignOfferContribution", 0),
        "Retailer Offer (£)":    v.get("RetailerOfferContribution", 0),
        # Vehicle details
        "Age (Years)":           v.get("Age"),
        "Mileage":               int(v.get("Mileage", 0)),
        "Fuel Type":             v.get("FuelType"),
        "Transmission":          v.get("TransmissionType"),
        "Colour":                v.get("Colour"),
        "Colour Group":          v.get("ColourGroup"),
        "Vehicle Source":        source.get("Description"),
        # Engine
        "Engine Badge":          engine.get("Badge"),
        "CC":                    engine.get("CubicCapacity", 0),
        "Cylinders":             engine.get("Cyclinders", 0),
        "Horsepower (BHP)":      float(engine.get("HorsePower", 0) or 0),
        "Power (kW)":            float(engine.get("Kw", 0) or 0),
        # Environmental
        "CO2 (g/km)":            v.get("Emission", 0),
        "Electric Range (mi)":   v.get("ElectricRange", 0),
        # Dealer
        "Dealer Name":           retailer.get("Description"),
        "Dealer City":           retailer.get("City"),
        "Dealer Postcode":       retailer.get("Postcode"),
        "Dealer Phone":          retailer.get("DisplayPhoneNumberUsed", retailer.get("Phone")),
        "Dealer Email":          retailer.get("Email"),
        "Dealer Website":        retailer.get("Website"),
        "Market Area":           retailer.get("MarketAreaName"),
        "Dealer Latitude":       retailer.get("Latitude"),
        "Dealer Longitude":      retailer.get("Longitude"),
        # Media
        "Image URL":             v.get("Media", {}).get("MainImageUrl"),
    }


def extract_new(v: dict) -> dict:
    engine = v.get("Engine", {})
    retailer = v.get("Retailer", {})
    brand = v.get("Brand", {})

    # Extract active offer descriptions
    offers = v.get("Offers", {}).get("Combined", {})
    combined = offers.get("Combined", [])
    offer_desc = combined[0].get("Value", "") if combined else ""
    offer_from = combined[0].get("From", "") if combined else ""
    offer_to = combined[0].get("To", "") if combined else ""

    # Extract finance APR pill if present
    finance_msgs = offers.get("FinanceOfferMessages", [])
    apr_label = finance_msgs[0].get("Label", "") if finance_msgs else ""

    return {
        # Identity
        "Listing ID":            v.get("Id"),
        "VIN":                   v.get("Vin"),
        "Commission Number":     v.get("CommissionNumber"),
        "Marketing Code":        v.get("MarketingCode"),
        "Model Year Code":       v.get("FullModelYearCode"),
        # Model
        "Brand":                 brand.get("Description", v.get("VehicleClass", "")),
        "Model":                 v.get("Model"),
        "Variant":               v.get("Description"),
        "Trim":                  v.get("ShortDescription"),
        "Body Style":            v.get("BodyStyle"),
        # Pricing
        "Actual Price (£)":      v.get("ActualPrice", 0),
        "Retail Price inc VAT (£)": v.get("RetailPriceIncVAT", 0),
        "OTR (£)":               v.get("OTR", 0),
        "P11D (£)":              v.get("P11D", 0),
        "Total Offer Value (£)": v.get("TotalOfferValue", 0),
        "Has Offer":             "Yes" if v.get("HasPromotionalOffer") else "No",
        "Offer Description":     offer_desc,
        "Offer From":            offer_from,
        "Offer To":              offer_to,
        "Finance APR":           apr_label,
        "Offer Expiry Date":     v.get("OfferExpiryDate", ""),
        "Campaign Offer (£)":    v.get("CampaignOfferContribution", 0),
        # Vehicle details
        "Fuel Type":             v.get("FuelType"),
        "Transmission":          v.get("TransmissionType"),
        "Colour":                v.get("Colour"),
        "Colour Group":          v.get("ColourGroup"),
        "Port Arrival Date":     v.get("PortArrivalDate", ""),
        "Delivery Time (days)":  v.get("DeliveryTime", 0),
        "Is Display Stock":      "Yes" if v.get("IsDisplayStock") else "No",
        "Under Offer":           "Yes" if v.get("UnderOffer") else "No",
        "Is Sellable":           "Yes" if v.get("IsSellable") else "No",
        # Engine
        "Engine Badge":          engine.get("Badge"),
        "CC":                    engine.get("CubicCapacity", 0),
        "Cylinders":             engine.get("Cyclinders", 0),
        "Horsepower (BHP)":      float(engine.get("HorsePower", 0) or 0),
        "Power (kW)":            float(engine.get("Kw", 0) or 0),
        # Environmental
        "CO2 (g/km)":            v.get("Emission", 0),
        # Dealer
        "Dealer Name":           retailer.get("Description"),
        "Dealer City":           retailer.get("City"),
        "Dealer Postcode":       retailer.get("Postcode"),
        "Dealer Phone":          retailer.get("DisplayPhoneNumberNew", retailer.get("Phone")),
        "Dealer Email":          retailer.get("Email"),
        "Dealer Website":        retailer.get("Website"),
        "Market Area":           retailer.get("MarketAreaName"),
        "Dealer Latitude":       retailer.get("Latitude"),
        "Dealer Longitude":      retailer.get("Longitude"),
        # Media
        "Image URL":             v.get("Media", {}).get("MainImageUrl"),
    }


# ── Styles ────────────────────────────────────────────────────────────────────
DARK = "1A1A1A"
LIGHT_GREY = "F4F4F4"
WHITE = "FFFFFF"

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def hdr(cell):
    cell.fill = PatternFill("solid", fgColor=DARK)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=9)
    cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def dat(cell, alt=False, align="left", fmt=None):
    cell.fill = PatternFill("solid", fgColor=LIGHT_GREY if alt else WHITE)
    cell.font = Font(name="Arial", size=9)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt


# ── Column width maps ─────────────────────────────────────────────────────────
USED_WIDTHS = {
    "Listing ID": 10, "VIN": 18, "Registration": 13, "Registration Date": 14,
    "Brand": 10, "Model": 18, "Variant": 20, "Trim": 22, "Body Style": 11,
    "Marketing Code": 12, "Price (£)": 12, "Total Offer Value (£)": 14,
    "Has Offer": 9, "Campaign Offer (£)": 14, "Retailer Offer (£)": 14,
    "Age (Years)": 9, "Mileage": 11, "Fuel Type": 10, "Transmission": 12,
    "Colour": 22, "Colour Group": 12, "Vehicle Source": 18,
    "Engine Badge": 10, "CC": 7, "Cylinders": 9,
    "Horsepower (BHP)": 13, "Power (kW)": 10,
    "CO2 (g/km)": 10, "Electric Range (mi)": 15,
    "Dealer Name": 28, "Dealer City": 16, "Dealer Postcode": 13,
    "Dealer Phone": 16, "Dealer Email": 30, "Dealer Website": 36,
    "Market Area": 24, "Dealer Latitude": 13, "Dealer Longitude": 14,
    "Image URL": 60,
}

NEW_WIDTHS = {
    "Listing ID": 10, "VIN": 18, "Commission Number": 16, "Marketing Code": 12,
    "Model Year Code": 12, "Brand": 10, "Model": 18, "Variant": 22, "Trim": 24,
    "Body Style": 11, "Actual Price (£)": 14, "Retail Price inc VAT (£)": 18,
    "OTR (£)": 12, "P11D (£)": 12, "Total Offer Value (£)": 14,
    "Has Offer": 9, "Offer Description": 20, "Offer From": 12, "Offer To": 12,
    "Finance APR": 12, "Offer Expiry Date": 14, "Campaign Offer (£)": 14,
    "Fuel Type": 10, "Transmission": 12, "Colour": 22, "Colour Group": 12,
    "Port Arrival Date": 16, "Delivery Time (days)": 16,
    "Is Display Stock": 13, "Under Offer": 10, "Is Sellable": 10,
    "Engine Badge": 10, "CC": 7, "Cylinders": 9,
    "Horsepower (BHP)": 13, "Power (kW)": 10, "CO2 (g/km)": 10,
    "Dealer Name": 28, "Dealer City": 16, "Dealer Postcode": 13,
    "Dealer Phone": 16, "Dealer Email": 30, "Dealer Website": 36,
    "Market Area": 24, "Dealer Latitude": 13, "Dealer Longitude": 14,
    "Image URL": 60,
}

# ── GBP / numeric columns ─────────────────────────────────────────────────────
USED_GBP = {"Price (£)", "Total Offer Value (£)",
            "Campaign Offer (£)", "Retailer Offer (£)"}
USED_NUM = {"Mileage", "CC"}
USED_RIGHT = USED_GBP | USED_NUM | {
    "Age (Years)", "Horsepower (BHP)", "Power (kW)", "CO2 (g/km)",
    "Electric Range (mi)", "Cylinders", "Dealer Latitude", "Dealer Longitude"
}

NEW_GBP = {
    "Actual Price (£)", "Retail Price inc VAT (£)", "OTR (£)",
    "P11D (£)", "Total Offer Value (£)", "Campaign Offer (£)"
}
NEW_NUM = {"CC", "Delivery Time (days)"}
NEW_RIGHT = NEW_GBP | NEW_NUM | {
    "Horsepower (BHP)", "Power (kW)", "CO2 (g/km)",
    "Cylinders", "Dealer Latitude", "Dealer Longitude"
}


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if not os.path.exists(INPUT_FILE):
        print(f"❌ File not found: {INPUT_FILE}")
        print(f"   Run scrape_{'new' if mode == 'new' else 'used'}.py first.")
        return

    print(f"📂 Loading {INPUT_FILE}...")
    with open(INPUT_FILE, encoding="utf-8") as f:
        raw = json.load(f)

    extract_fn = extract_new if mode == "new" else extract_used
    GBP_COLS = NEW_GBP if mode == "new" else USED_GBP
    NUM_COLS = NEW_NUM if mode == "new" else USED_NUM
    RIGHT_COLS = NEW_RIGHT if mode == "new" else USED_RIGHT
    WIDTHS = NEW_WIDTHS if mode == "new" else USED_WIDTHS

    seen, rows = set(), []
    for v in raw:
        uid = v.get("Id")
        if uid not in seen:
            seen.add(uid)
            rows.append(extract_fn(v))

    print(f"✅ {len(rows):,} unique vehicles extracted")

    os.makedirs("output", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_LABEL
    ws.sheet_view.showGridLines = False

    headers = list(rows[0].keys())

    ws.row_dimensions[1].height = 32
    for ci, h in enumerate(headers, 1):
        hdr(ws.cell(row=1, column=ci, value=h))

    gbp_idx = {headers.index(c) + 1 for c in GBP_COLS if c in headers}
    num_idx = {headers.index(c) + 1 for c in NUM_COLS if c in headers}
    right_idx = {headers.index(c) + 1 for c in RIGHT_COLS if c in headers}

    for ri, row in enumerate(rows, 2):
        alt = ri % 2 == 0
        ws.row_dimensions[ri].height = 16
        for ci, key in enumerate(headers, 1):
            val = row[key]
            algn = "right" if ci in right_idx else "left"
            fmt = "£#,##0" if ci in gbp_idx else (
                "#,##0" if ci in num_idx else None)
            dat(ws.cell(row=ri, column=ci, value=val),
                alt=alt, align=algn, fmt=fmt)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"

    for ci, key in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = WIDTHS.get(key, 14)

    wb.save(OUTPUT_FILE)
    print(f"💾 Saved → {OUTPUT_FILE}")
    print(f"\n📊 {len(headers)} columns written for {mode} cars")


if __name__ == "__main__":
    main()
