"""
to_powerbi.py
─────────────
Reads both used and new car JSON files and produces a single
Power BI-ready Excel workbook with:

  Sheet 1 - All Vehicles   : unified dataset (used + new) with Vehicle_Type column
  Sheet 2 - Used Vehicles   : used-only with all used-specific fields
  Sheet 3 - New Vehicles    : new-only with all new-specific fields
  Sheet 4 - Dealers         : unique dealer reference table

HOW TO USE:
  python to_powerbi.py

Requirements:
  pip install openpyxl
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

USED_FILE = "data/used/mercedes_used_cars_FULL.json"
NEW_FILE = "data/new/mercedes_new_cars_FULL.json"
OUTPUT_FILE = "output/mercedes_powerbi.xlsx"

# ── Extraction ────────────────────────────────────────────────────────────────


def extract_unified(v, vehicle_type):
    """Shared fields that exist in both used and new — for the combined sheet."""
    engine = v.get("Engine", {})
    retailer = v.get("Retailer", {})
    brand = v.get("Brand", {})
    return {
        "Vehicle_Type":          vehicle_type,
        "Listing_ID":            v.get("Id"),
        "VIN":                   v.get("Vin"),
        "Brand":                 brand.get("Description", v.get("VehicleClass", "")),
        "Model":                 v.get("Model"),
        "Variant":               v.get("Description"),
        "Trim":                  v.get("ShortDescription"),
        "Body_Style":            v.get("BodyStyle"),
        "Marketing_Code":        v.get("MarketingCode"),
        "Fuel_Type":             v.get("FuelType"),
        "Transmission":          v.get("TransmissionType"),
        "Colour_Group":          v.get("ColourGroup"),
        "Colour":                v.get("Colour"),
        "Horsepower_BHP":        float(engine.get("HorsePower", 0) or 0),
        "Power_kW":              float(engine.get("Kw", 0) or 0),
        "Engine_CC":             engine.get("CubicCapacity", 0),
        "Cylinders":             engine.get("Cyclinders", 0),
        "CO2_g_km":              v.get("Emission", 0),
        "Electric_Range_mi":     v.get("ElectricRange", 0),
        "Actual_Price_GBP":      v.get("ActualPrice", 0),
        "Total_Offer_Value_GBP": v.get("TotalOfferValue", 0),
        "Has_Offer":             1 if v.get("HasPromotionalOffer") else 0,
        "Campaign_Offer_GBP":    v.get("CampaignOfferContribution", 0),
        "Retailer_Offer_GBP":    v.get("RetailerOfferContribution", 0),
        # Used-only fields (blank for new)
        "Age_Years":             v.get("Age", None),
        "Mileage":               int(v.get("Mileage", 0)) if v.get("Mileage") else None,
        "Registration_Date":     v.get("RegistrationDate", ""),
        "Vehicle_Source":        v.get("VehicleSource", {}).get("Description", ""),
        # New-only fields (blank for used)
        "OTR_GBP":               v.get("OTR", None),
        "P11D_GBP":              v.get("P11D", None),
        "Retail_Price_incVAT":   v.get("RetailPriceIncVAT", None),
        "Port_Arrival_Date":     v.get("PortArrivalDate", ""),
        "Delivery_Days":         v.get("DeliveryTime", None),
        "Offer_Expiry_Date":     v.get("OfferExpiryDate", ""),
        "Is_Display_Stock":      1 if v.get("IsDisplayStock") else 0,
        "Under_Offer":           1 if v.get("UnderOffer") else 0,
        # Dealer
        "Dealer_ID":             retailer.get("Id"),
        "Dealer_Name":           retailer.get("Description"),
        "Dealer_City":           retailer.get("City"),
        "Dealer_Postcode":       retailer.get("Postcode"),
        "Market_Area":           retailer.get("MarketAreaName"),
        "Dealer_Latitude":       retailer.get("Latitude"),
        "Dealer_Longitude":      retailer.get("Longitude"),
    }


def extract_used(v):
    engine = v.get("Engine", {})
    retailer = v.get("Retailer", {})
    source = v.get("VehicleSource", {})
    brand = v.get("Brand", {})
    return {
        "Listing_ID":            v.get("Id"),
        "VIN":                   v.get("Vin"),
        "Registration":          v.get("RegistrationNumber"),
        "Registration_Date":     v.get("RegistrationDate"),
        "Brand":                 brand.get("Description", v.get("VehicleClass", "")),
        "Model":                 v.get("Model"),
        "Variant":               v.get("Description"),
        "Trim":                  v.get("ShortDescription"),
        "Body_Style":            v.get("BodyStyle"),
        "Marketing_Code":        v.get("MarketingCode"),
        "Actual_Price_GBP":      v.get("ActualPrice", 0),
        "Total_Offer_Value_GBP": v.get("TotalOfferValue", 0),
        "Has_Offer":             1 if v.get("HasPromotionalOffer") else 0,
        "Campaign_Offer_GBP":    v.get("CampaignOfferContribution", 0),
        "Retailer_Offer_GBP":    v.get("RetailerOfferContribution", 0),
        "Age_Years":             v.get("Age"),
        "Mileage":               int(v.get("Mileage", 0)),
        "Fuel_Type":             v.get("FuelType"),
        "Transmission":          v.get("TransmissionType"),
        "Colour":                v.get("Colour"),
        "Colour_Group":          v.get("ColourGroup"),
        "Vehicle_Source":        source.get("Description"),
        "Engine_Badge":          engine.get("Badge"),
        "Engine_CC":             engine.get("CubicCapacity", 0),
        "Cylinders":             engine.get("Cyclinders", 0),
        "Horsepower_BHP":        float(engine.get("HorsePower", 0) or 0),
        "Power_kW":              float(engine.get("Kw", 0) or 0),
        "CO2_g_km":              v.get("Emission", 0),
        "Electric_Range_mi":     v.get("ElectricRange", 0),
        "Dealer_ID":             retailer.get("Id"),
        "Dealer_Name":           retailer.get("Description"),
        "Dealer_City":           retailer.get("City"),
        "Dealer_Postcode":       retailer.get("Postcode"),
        "Dealer_Phone":          retailer.get("DisplayPhoneNumberUsed", retailer.get("Phone")),
        "Dealer_Email":          retailer.get("Email"),
        "Dealer_Website":        retailer.get("Website"),
        "Market_Area":           retailer.get("MarketAreaName"),
        "Dealer_Latitude":       retailer.get("Latitude"),
        "Dealer_Longitude":      retailer.get("Longitude"),
        "Image_URL":             v.get("Media", {}).get("MainImageUrl"),
    }


def extract_new(v):
    engine = v.get("Engine", {})
    retailer = v.get("Retailer", {})
    brand = v.get("Brand", {})
    offers = v.get("Offers", {}).get("Combined", {})
    combined = offers.get("Combined", [])
    offer_desc = combined[0].get("Value", "") if combined else ""
    offer_from = combined[0].get("From", "") if combined else ""
    offer_to = combined[0].get("To", "") if combined else ""
    fin_msgs = offers.get("FinanceOfferMessages", [])
    apr_label = fin_msgs[0].get("Label", "") if fin_msgs else ""
    return {
        "Listing_ID":               v.get("Id"),
        "VIN":                      v.get("Vin"),
        "Commission_Number":        v.get("CommissionNumber"),
        "Marketing_Code":           v.get("MarketingCode"),
        "Model_Year_Code":          v.get("FullModelYearCode"),
        "Brand":                    brand.get("Description", v.get("VehicleClass", "")),
        "Model":                    v.get("Model"),
        "Variant":                  v.get("Description"),
        "Trim":                     v.get("ShortDescription"),
        "Body_Style":               v.get("BodyStyle"),
        "Actual_Price_GBP":         v.get("ActualPrice", 0),
        "Retail_Price_incVAT_GBP":  v.get("RetailPriceIncVAT", 0),
        "OTR_GBP":                  v.get("OTR", 0),
        "P11D_GBP":                 v.get("P11D", 0),
        "Total_Offer_Value_GBP":    v.get("TotalOfferValue", 0),
        "Has_Offer":                1 if v.get("HasPromotionalOffer") else 0,
        "Offer_Description":        offer_desc,
        "Offer_From":               offer_from,
        "Offer_To":                 offer_to,
        "Finance_APR":              apr_label,
        "Offer_Expiry_Date":        v.get("OfferExpiryDate", ""),
        "Campaign_Offer_GBP":       v.get("CampaignOfferContribution", 0),
        "Fuel_Type":                v.get("FuelType"),
        "Transmission":             v.get("TransmissionType"),
        "Colour":                   v.get("Colour"),
        "Colour_Group":             v.get("ColourGroup"),
        "Port_Arrival_Date":        v.get("PortArrivalDate", ""),
        "Delivery_Days":            v.get("DeliveryTime", 0),
        "Is_Display_Stock":         1 if v.get("IsDisplayStock") else 0,
        "Under_Offer":              1 if v.get("UnderOffer") else 0,
        "Is_Sellable":              1 if v.get("IsSellable") else 0,
        "Engine_Badge":             engine.get("Badge"),
        "Engine_CC":                engine.get("CubicCapacity", 0),
        "Cylinders":                engine.get("Cyclinders", 0),
        "Horsepower_BHP":           float(engine.get("HorsePower", 0) or 0),
        "Power_kW":                 float(engine.get("Kw", 0) or 0),
        "CO2_g_km":                 v.get("Emission", 0),
        "Dealer_ID":                retailer.get("Id"),
        "Dealer_Name":              retailer.get("Description"),
        "Dealer_City":              retailer.get("City"),
        "Dealer_Postcode":          retailer.get("Postcode"),
        "Dealer_Phone":             retailer.get("DisplayPhoneNumberNew", retailer.get("Phone")),
        "Dealer_Email":             retailer.get("Email"),
        "Dealer_Website":           retailer.get("Website"),
        "Market_Area":              retailer.get("MarketAreaName"),
        "Dealer_Latitude":          retailer.get("Latitude"),
        "Dealer_Longitude":         retailer.get("Longitude"),
        "Image_URL":                v.get("Media", {}).get("MainImageUrl"),
    }


def extract_dealer(retailer):
    return {
        "Dealer_ID":        retailer.get("Id"),
        "Dealer_Name":      retailer.get("Description"),
        "Street":           retailer.get("Street"),
        "City":             retailer.get("City"),
        "Postcode":         retailer.get("Postcode"),
        "Phone":            retailer.get("Phone"),
        "Email":            retailer.get("Email"),
        "Website":          retailer.get("Website"),
        "Latitude":         retailer.get("Latitude"),
        "Longitude":        retailer.get("Longitude"),
        "Market_Area":      retailer.get("MarketAreaName"),
        "Retailer_Group":   retailer.get("RetailerGroupName"),
        "Is_New_Retailer":  1 if retailer.get("IsNewCarRetailer") else 0,
        "Is_Used_Retailer": 1 if retailer.get("IsUsedCarRetailer") else 0,
    }


# ── Styles ────────────────────────────────────────────────────────────────────
DARK = "1A1A1A"
WHITE = "FFFFFF"
LIGHT = "F4F4F4"
thin = Side(style="thin", color="CCCCCC")
BDR = Border(left=thin, right=thin, top=thin, bottom=thin)


def hdr(cell, color=DARK):
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=9)
    cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)
    cell.border = BDR


def dat(cell, alt=False):
    cell.fill = PatternFill("solid", fgColor=LIGHT if alt else WHITE)
    cell.font = Font(name="Arial", size=9)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BDR


def write_sheet(ws, rows, header_color=DARK, col_width=18):
    if not rows:
        return
    ws.sheet_view.showGridLines = False
    headers = list(rows[0].keys())
    ws.row_dimensions[1].height = 28
    for ci, h in enumerate(headers, 1):
        hdr(ws.cell(row=1, column=ci, value=h), color=header_color)
        ws.column_dimensions[get_column_letter(ci)].width = col_width
    for ri, row in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 15
        for ci, key in enumerate(headers, 1):
            dat(ws.cell(row=ri, column=ci, value=row[key]), alt=ri % 2 == 0)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    used_raw, new_raw = [], []

    if os.path.exists(USED_FILE):
        print(f"📂 Loading used cars...")
        with open(USED_FILE, encoding="utf-8") as f:
            used_raw = json.load(f)
        print(f"   {len(used_raw):,} records")
    else:
        print(f"⚠️  {USED_FILE} not found — skipping used cars")

    if os.path.exists(NEW_FILE):
        print(f"📂 Loading new cars...")
        with open(NEW_FILE, encoding="utf-8") as f:
            new_raw = json.load(f)
        print(f"   {len(new_raw):,} records")
    else:
        print(f"⚠️  {NEW_FILE} not found — skipping new cars")

    if not used_raw and not new_raw:
        print("❌ No data found. Run scrapers first.")
        return

    # Deduplicate
    seen_used, used_rows, used_unified = set(), [], []
    for v in used_raw:
        uid = v.get("Id")
        if uid not in seen_used:
            seen_used.add(uid)
            used_rows.append(extract_used(v))
            used_unified.append(extract_unified(v, "Used"))

    seen_new, new_rows, new_unified = set(), [], []
    for v in new_raw:
        uid = v.get("Id")
        if uid not in seen_new:
            seen_new.add(uid)
            new_rows.append(extract_new(v))
            new_unified.append(extract_unified(v, "New"))

    all_unified = used_unified + new_unified

    # Dealer reference table
    seen_dealers, dealer_rows = set(), []
    for v in used_raw + new_raw:
        r = v.get("Retailer", {})
        did = r.get("Id")
        if did and did not in seen_dealers:
            seen_dealers.add(did)
            dealer_rows.append(extract_dealer(r))

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)

    ws_all = wb.create_sheet("All_Vehicles")
    write_sheet(ws_all, all_unified, header_color="1A1A1A", col_width=20)

    if used_rows:
        ws_used = wb.create_sheet("Used_Vehicles")
        write_sheet(ws_used, used_rows, header_color="1A3A5C", col_width=18)

    if new_rows:
        ws_new = wb.create_sheet("New_Vehicles")
        write_sheet(ws_new, new_rows, header_color="1A4A2A", col_width=18)

    ws_dealers = wb.create_sheet("Dealers")
    write_sheet(ws_dealers, dealer_rows, header_color="4A1A1A", col_width=22)

    os.makedirs("output", exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Power BI file saved → {OUTPUT_FILE}")
    print(f"   All_Vehicles sheet : {len(all_unified):,} rows")
    print(f"   Used_Vehicles sheet: {len(used_rows):,} rows")
    print(f"   New_Vehicles sheet : {len(new_rows):,} rows")
    print(f"   Dealers sheet      : {len(dealer_rows):,} rows")


if __name__ == "__main__":
    main()
